"""
Given the Excel file, return filtered variants in a dictionary

A variant can have multiple entries, so need to be careful
"""

import _io
import logging
from typing import List
from openpyxl import load_workbook
from tqdm import tqdm

logger = logging.getLogger()

CONF_CLASSES = {
    'assoc_resistance': '1) Assoc w R',
    'assoc_resistance_interim': '2) Assoc w R - Interim',
    'uncert_signif': '3) Uncertain significance',
    'no_assoc_interim': '4) Not assoc w R - Interim',
    'no_assoc': '5) Not assoc w R',
    'combo': 'combo'
}


def parse_file(filehandle: _io.BytesIO, this_filter: str = 'assoc_resistance') -> List[dict]:
    """

    :param filehandle: binary file-like object
    :param this_filter: confidence class to filter on
    :return: dictionary
    """
    logger.info("Processing Excel file...")

    if this_filter == 'all':
        conf_class = list(CONF_CLASSES.values())
    else:
        conf_class = [CONF_CLASSES[this_filter]]
    logger.info("Opening the Excel file...")
    wb = load_workbook(filename=filehandle, read_only=True)
    sh = wb['Mutation_catalogue']
    val_iter = sh.values
    header = next(val_iter)
    drug_ix = header.index("drug")
    var_ix = header.index("variant (common_name)")
    genome_pos_ix = header.index("Genome position")
    conf_class_ix = header.index("FINAL CONFIDENCE GRADING")
    logger.info("Processing the rows...")
    logger.info("Progress....")
    data = [{'drug': row[drug_ix],
             'var': row[var_ix],
             'genome_pos': row[genome_pos_ix],
             'confidence': row[conf_class_ix]} for row in tqdm(val_iter, total=sh.max_row - 1) if
            row[conf_class_ix] in conf_class]
    logger.info(f"Kept {len(data)} records...")
    return data
